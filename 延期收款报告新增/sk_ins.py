#延期收款报告调整 web自动化程序
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoAlertPresentException,TimeoutException,NoSuchElementException
from selenium.webdriver import ActionChains
from openpyxl import load_workbook
import keyboard
import sys
import os


#点击收款报告新增
def yqskbgxz_click(wd):
    #点击数据申报
    element_1 = wd.find_element(By.XPATH,'//*[@id="646"]')
    element_1.click()
    #点击货物贸易
    element_1 = wd.find_element(By.XPATH,'//*[@id="13073c56-16712888ba7-81d76f67e9fa75bc16134518e029baa9"]')
    element_1.click()
     #点击贸易信贷与融资报告
    element_1 = wd.find_element(By.XPATH,'//*[@id="res13073c56-1671288960b-81d76f67e9fa75bc16134518e029baa9"]')
    element_1.click()
    #点击延期收款报告
    element_1 = wd.find_element(By.XPATH,'//*[@id="res13073c56-16712889db5-81d76f67e9fa75bc16134518e029baa9"]')
    element_1.click()
    #点击延期收款新增
    element_1 = wd.find_element(By.XPATH,'//*[@id="res13073c56-1671288a3a3-81d76f67e9fa75bc16134518e029baa9"]')
    element_1.click()

#保存文件   
def save_sheet(workbook):
    print('save')
    path = os.path.dirname(os.path.realpath(__file__))
    print(path)
    workbook.save(path + '\sk_xinzeng_complete.xlsx')

#异常后重新调用main_work
def main_work_again(ah,ws,readws,i):
    ah.wd.switch_to.default_content()
    ah.wd.close()
    ah.widowChang('国家外汇管理')
    ah.wd.switch_to.frame('myframe')
    main_work(ah,ws,readws,i)
    
#打开调整画面后的主要程序
def main_work(ah,ws,readws,i):
    try:
        #点击第一个需要改的单子
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="page-0"]/div/table/tbody/tr/td/input')
        element_1.click()
        #点击新增
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_2"]/nobr/span[2]')
        element_1.click()
        mainWindow = ah.wd.current_window_handle
        ah.widowChang('延期收款报告')
        ah.wd.maximize_window()
        #切换frame
        ah.wd.switch_to.frame('containerIframe')
        ah.load_complete('//*[@id="unieap-loading"]')
        #复制金额
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_InlineEditBox_4"]/div/div[2]')
        text = element_1.text
        print(text)
        #点击新增
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_0"]/nobr/span[2]')
        element_1.click()
        #写入日期
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_DateTextBox_0"]/div/div[3]/input')
        ActionChains(ah.wd).double_click(element_1).perform()
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_DateTextBox_0"]/div/div[3]/input')
        print(readws.cell(i,4).value)
        element_1.send_keys(readws.cell(i,4).value)

         #写入金额
        #element_1 = ah.wd.find_element(By.XPATH,'//*[@id="page-0"]/div[2]/table/tbody/tr/td[5]')
        #ActionChains(ah.wd).double_click(element_1).perform()
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_NumberTextBox_0"]/div/div[3]/input')
        element_1.send_keys(text)

        #提交
        element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_2"]/nobr/span[2]')
        element_1.click()
        #确定
        ah.confirm_click()
        #等待数据装载
        ah.load_complete('//*[@id="unieap-loading"]')
        #判断结果
        elements = ah.wd.find_elements(By.XPATH,'//*[@id="unieap_dialog_Dialog_1"]/div[2]/div[3]/div[1]/div/table/tbody/tr/td/table/tbody/tr/td[2]')
        for element in elements:
            print(element.text)
            if(element.text == '报告新增成功！' or element.text == "删除明细记录时获取报告序号失败！"):
                ws.cell(i, 6, '报告调整成功！')
                ws.cell(i, 5, 'success')
            else:
                ws.cell(i, 6, element.text)
                ws.cell(i, 5, 'fail')     
        ah.wd.switch_to.default_content()
        ah.wd.close()
        ah.wd.switch_to.window(mainWindow)
        ah.wd.switch_to.frame('myframe')
    except Exception as e:
        except_type, except_value, except_traceback = sys.exc_info()
        except_file = os.path.split(except_traceback.tb_frame.f_code.co_filename)[1]
        exc_dict = {
            "报错类型": except_type,
            "报错信息": except_value,
            "报错文件": except_file,
            "报错行数": except_traceback.tb_lineno,
        }
        print(exc_dict)
        main_work_again(ah,ws,readws,i)
    else:
        print("mainwork is no problem")
    finally:
        print("mainwork_done")

#循环处理每行数据        
def excl_job(ah):
    try:
        #切换frame
        ah.wd.switch_to.frame('myframe')
        path = os.path.dirname(os.path.realpath(__file__))
        print(path)
        workbook = load_workbook(path + '\sk_xinzeng.xlsx',data_only=True)
        readws = workbook.active
        workbooknew = load_workbook(path + '\sk_xinzeng.xlsx')
        ws = workbooknew.active
        keyboard.add_hotkey('ctrl+alt', save_sheet,args=(workbooknew,))
        save_sheet(workbooknew)
        i = 0
        for row in readws.rows: # 循环逐行打印
            i = i + 1
            if(i == 1):
                continue
            if(readws.cell(i,5).value == 'success' or readws.cell(i,5).value == 'continue'):
                continue
            ws.cell(i,5,' ')
            if(readws.cell(i,2).value != ' ' and readws.cell(i,2).value is not None):
                print(readws.cell(i,2).value)
                #重置
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_1"]/nobr/span[2]')
                element_1.click()
                #报关单赋值
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="entryId"]/div/div[3]/input')
                element_1.send_keys(readws.cell(i,2).value)
                #查询
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_0"]/nobr/span[2]')
                element_1.click()
                #判断是否找到数据，load
                ah.load_complete('//*[@id="unieap-loading"]')
                #点单选选择项
                flag = ah.is_element_exist('//*[@id="page-0"]/div/table/tbody/tr/td/input')
                if(flag == True):
                    main_work(ah,ws,readws,i)
                    continue
                else:
                    #未找到数据处理
                    ah.confirm_click()
                    ws.cell(i,6,'报关单未找到')
            if(readws.cell(i,1).value != ' ' and readws.cell(i,1).value is not None):
                print(readws.cell(i,1).value)
                #重置
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_1"]/nobr/span[2]')
                element_1.click()
                #合同赋值
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="contractno"]/div/div[3]/input')
                element_1.send_keys(readws.cell(i,1).value)
                #查询
                element_1 = ah.wd.find_element(By.XPATH,'//*[@id="unieap_form_Button_0"]/nobr/span[2]')
                element_1.click()
                #判断是否找到数据，load
                ah.load_complete('//*[@id="unieap-loading"]')
                #点单选选择项
                flag = ah.is_element_exist('//*[@id="page-0"]/div/table/tbody/tr/td/input')
                if(flag == True):
                    main_work(ah,ws,readws,i)
                else:
                    #未找到数据处理
                    ah.confirm_click()
                    ws.cell(i, 6, '合同未找到')
        #切换到默认frame
        ah.wd.switch_to.default_content()
    except Exception as e:
        except_type, except_value, except_traceback = sys.exc_info()
        except_file = os.path.split(except_traceback.tb_frame.f_code.co_filename)[1]
        exc_dict = {
            "报错类型": except_type,
            "报错信息": except_value,
            "报错文件": except_file,
            "报错行数": except_traceback.tb_lineno,
        }
        print(exc_dict)
    else:
        print("the code is no problem")
    finally:
        save_sheet(workbooknew)
        print("this is finally code,i'm running")    
            

if __name__ == '__main__':
    up_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    sys.path.append(up_path)
    from autoHelper import autoHelper
    # 创建 WebDriver 对象，指明使用chrome浏览器驱动
    ah = autoHelper('http://zwfw.safe.gov.cn/asone/servlet/UniLoginServlet?COLLCC=1430122602&')
    #wd = webdriver.Chrome()
    #最大化窗口，可有可无
    #wd.maximize_window()
    ah.wd.maximize_window()
    #所有查找元素等待五秒轮询
    #wd.implicitly_wait(5)
    ah.wd.implicitly_wait(5)
    #load(wd)
    ah.load()
    #点击延期收款报告
    yqskbgxz_click(ah.wd)
    #循环处理EXCL
    excl_job(ah)


